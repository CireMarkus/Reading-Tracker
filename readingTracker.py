#Packages to manipulate excel files
from tkinter.constants import BOTH, END
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
#Book class
from Book import Book
#graphical imports 
import tkinter









def populate(currentBooks):
    for i in range(len(shelf)):
        title = tkinter.Label(master=currentBooks, text=str(shelf[i].getTitle()))
        title.grid(row= i+1, column = 0)

        currentPage = tkinter.Entry(master=currentBooks)
        currentPage.insert(END,str(shelf[i].getPagesRead()))
        currentPage.grid(row=i+1, column=1)

        percentCompleted = tkinter.Label(master=currentBooks, text=str(shelf[i].percentCompleted())+'%')
        percentCompleted.grid(row=i+1, column=2)
        
        completed = tkinter.Label(master=currentBooks, text=str(shelf[i].getDateCompleted()))
        completed.grid(row=i+1, column=3)

def updatePages():
    for i in range(len(shelf)):
        #sprint(currentBooks.grid_slaves(row = i+1, column = 1 )[0].get() )
        shelf[i].updatePagesRead(int(currentBooks.grid_slaves(row = i+1, column = 1 )[0].get()))
    populate(currentBooks)
    wb = Workbook()
    wb = load_workbook('Bookstoread.xlsx')
    ws = wb.active
    for i in range(0,len(shelf)):
        ws['B'+ str(i+2)] = shelf[i].getPagesRead()
        ws['D' + str(i+2)] = str(shelf[i].getCompleted())
        ws['E' + str(i+2)] = str(shelf[i].getDateCompleted())
    wb.save('Bookstoread.xlsx')


#Main program below
wb = Workbook()
wb = load_workbook('Bookstoread.xlsx')
ws = wb.active
shelf = []
for row in ws.iter_rows(min_row=2, values_only = True):
    shelf.append(Book(row[0], row[1], row[2], row[3], row[4]))




#Tikinter goes below here
window = tkinter.Tk()
window.geometry("800x500")
window.title("Reading Tracker")


#Frame that shows the books that are currently being read. 
currentBooks = tkinter.Frame(master=window)
currentBooks.pack(fill=BOTH)
titleLabel = tkinter.Label(master=currentBooks, text="Title")
titleLabel.grid(row = 0, column=0)
currentPageLabel = tkinter.Label(master=currentBooks, text="Current Page")
currentPageLabel.grid(row=0,column=1)
percentCompletedLabel = tkinter.Label(master=currentBooks, text="Completed")
percentCompletedLabel.grid(row=0, column=2)
percentCompletedLabel = tkinter.Label(master=currentBooks, text="Date Completed")
percentCompletedLabel.grid(row=0, column=3)
populate(currentBooks)

savePages = tkinter.Button(master=currentBooks, height= 2, width = 10, command= updatePages, text='Save')
savePages.grid(row=len(shelf)+1, column = 4,)

window.mainloop()


